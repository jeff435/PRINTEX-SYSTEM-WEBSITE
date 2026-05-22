import openpyxl
import re
import json

wb = openpyxl.load_workbook(r"C:\Users\USER\Desktop\PRINTEX ENGINEERS LIMITED SPARE PARTS august 2025docx.xlsx", data_only=True)
sheet = wb.active

def clean_price(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip().replace("$", "").replace("Ksh", "").replace("KES", "").replace(",", "").replace("&", "")
    try:
        return float(val_str)
    except ValueError:
        return 0

def extract_sku(desc):
    patterns = [
        r'[A-Z0-9]{2}\.[0-9]{3}\.[0-9]{3,4}(?:/[0-9]{2})?[A-Z]?',
        r'[A-Z0-9][0-9]\.[0-9]{3}\.[0-9]{3,4}(?:/[0-9]{2})?[A-Z]?',
        r'[0-9]{2}\.[0-9]{3}\.[0-9]{3,4}(?:/[0-9]{2})?[A-Z]?',
        r'[A-Z]{2}\.[0-9]{3}\.[0-9]{3,4}[A-Z]?',
        r'[0-9]{3}\.[0-9]{4}\.[0-9]{1,4}[A-Z]?',
        r'[0-9]{5}\.[0-9]{3}',
        r'[0-9]{2}\.[0-9]{7}',
        r'[A-Z][0-9]\.[0-9]{3}/[0-9]{2}',
        r'F-[0-9]{5,6}(?:\.[0-9]{1,2})?',
        r'F[0-9]{1}\.[0-9]{3}\.[0-9]{3}(?:\.[0-9]{1,2})?',
        r'HFL[0-9A-Z]+',
        r'[0-9]{3}-[0-9]{3}-[0-9]{2}',
        r'[A-Z]{2}\.[0-9]{3}\.[0-9]{3}[A-Z]?',
        r'[A-Z][0-9]\.[0-9]{3}\.[0-9]{3}[A-Z]?',
        r'[0-9]{2}\.[0-9]{3}\.[0-9]{4}',
        r'WL9-P[0-9]{3}',
        r'D8251A',
        r'89A32',
    ]
    for p in patterns:
        m = re.findall(p, desc)
        if m:
            sku = m[0]
            cleaned_desc = desc.replace(sku, "", 1).strip()
            cleaned_desc = re.sub(r'\s+-\s+', ' ', cleaned_desc)
            cleaned_desc = re.sub(r'\s+', ' ', cleaned_desc)
            cleaned_desc = cleaned_desc.strip(" -/,")
            return sku, cleaned_desc
    return "", desc

# We start reading from row 3
current_cat = "A"
parts_db = []
start_id = 129  # Assign IDs sequentially from 129 to 436

for r in range(3, sheet.max_row + 1):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 6)]
    if not any(row_vals):
        continue
    
    row_val = row_vals[0]
    item_val = row_vals[1]
    qty_val = row_vals[2]
    ksh_val = row_vals[3]
    usd_val = row_vals[4]
    
    # Check if this row is a column header or category marker
    if item_val and str(item_val).strip().startswith("COLUMN"):
        current_cat = str(item_val).strip().split(" ")[-1].strip()
        continue
    
    if row_val and str(row_val).strip() in ["A", "B", "C", "D", "E", "G", "F", "J", "L", "K"]:
        current_cat = str(row_val).strip()
        
    if not item_val:
        continue
        
    desc_raw = str(item_val).strip()
    if desc_raw == "ITEM":
        continue
        
    sku, desc_cleaned = extract_sku(desc_raw)
    
    # Parse quantities and prices
    qty = 0
    if qty_val is not None:
        try:
            qty = int(clean_price(qty_val))
        except:
            qty = 0
            
    ksh = clean_price(ksh_val)
    usd = clean_price(usd_val)
    
    if ksh == 0 and usd > 0:
        ksh = usd * 130
        
    desc_cleaned = re.sub(r'\s+', ' ', desc_cleaned).strip()
    
    supplier = "Other"
    loc_desc = desc_cleaned.lower()
    if "heidelberg" in loc_desc or "heildelberg" in loc_desc:
        supplier = "Heidelberg"
    elif "muller" in loc_desc or "martini" in loc_desc or "hohner" in loc_desc:
        supplier = "Muller Martini"
    elif "polar" in loc_desc:
        supplier = "Polar"
    elif "stahl" in loc_desc:
        supplier = "Stahl"
    elif "festo" in loc_desc:
        supplier = "Festo"
    elif "smc" in loc_desc:
        supplier = "SMC"
    elif "technotrans" in loc_desc:
        supplier = "Technotrans"
    elif "weko" in loc_desc:
        supplier = "Weko"
        
    loc_map = {
        'A': 'Warehouse A1',
        'B': 'Warehouse B1',
        'C': 'Warehouse C1',
        'D': 'Warehouse D1',
        'E': 'Warehouse E1',
        'F': 'Warehouse F1',
        'G': 'Warehouse G1',
        'J': 'Warehouse J1',
        'K': 'Warehouse K1',
        'L': 'Warehouse L1',
    }
    location = loc_map.get(current_cat, "Warehouse G1")
    
    min_stock = 1
    if qty > 10:
        min_stock = 5
    elif qty > 5:
        min_stock = 2
        
    if not desc_cleaned or desc_cleaned.isdigit():
        desc_cleaned = f"Spare Part {sku if sku else start_id}"
        
    parts_db.append({
        "id": start_id,
        "category": current_cat,
        "partNum": sku if sku else f"SKU-{start_id:03d}",
        "desc": desc_cleaned,
        "stock": qty,
        "minStock": min_stock,
        "priceKsh": int(ksh),
        "supplier": supplier,
        "location": location
    })
    start_id += 1

js_code = f"""
window.printexBulkImportData = {json.dumps(parts_db)};

window.runPrintexBulkImport = async function() {{
  if (window._printexImportRunning) return;
  window._printexImportRunning = true;
  
  const toImport = window.printexBulkImportData;
  console.log('Starting bulk import of ' + toImport.length + ' parts...');
  window.showToast('Starting bulk import of ' + toImport.length + ' parts...', 'info');
  
  let added = 0;
  let updated = 0;
  
  for (const part of toImport) {{
    const existingIdx = window.parts.findIndex(p => 
      p.partNum === part.partNum || 
      (p.part_num && p.part_num === part.partNum) ||
      (p.desc === part.desc && p.category === part.category)
    );
    
    if (existingIdx !== -1) {{
      // Update existing
      const existing = window.parts[existingIdx];
      existing.stock = part.stock; // Update quantity
      existing.priceKsh = part.priceKsh; // Update price
      existing.supplier = part.supplier;
      existing.location = part.location;
      existing.desc = part.desc;
      existing.category = part.category;
      
      try {{
        await window.dbPut('parts', existing);
        updated++;
      }} catch (e) {{
        console.error('Failed to update part', part, e);
      }}
    }} else {{
      // Add new
      const newPart = {{ ...part, id: 'prt_' + Math.random().toString(36).substring(2, 9) + '_' + Date.now() }};
      try {{
        await window.dbPut('parts', newPart);
        window.parts.push(newPart);
        added++;
      }} catch (e) {{
        console.error('Failed to add part', part, e);
      }}
    }}
  }}
  
  window.showToast(`Import complete! Added ${{added}}, Updated ${{updated}} parts.`, 'success');
  console.log(`Import complete! Added ${{added}}, Updated ${{updated}} parts.`);
  
  if (typeof window.filterInventory === 'function') window.filterInventory();
  if (typeof window.renderDashboard === 'function') window.renderDashboard();
  
  window._printexImportRunning = false;
}};

// ── CRITICAL: Expose the parts list as window.DEFAULT_PARTS ──────────────────
window.DEFAULT_PARTS = window.printexBulkImportData;
"""

with open("src/modules/bulk_import.js", "w", encoding="utf-8") as f:
    f.write(js_code)

print(f"Generated bulk_import.js with {len(parts_db)} parts. Min ID: {parts_db[0]['id']}, Max ID: {parts_db[-1]['id']}")
